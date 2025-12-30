from dataclasses import dataclass, field
from typing import Dict, List, Optional
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import config
import re
from normalization_config import normalize_station_name

@dataclass
class Intern:
    name: str
    start_date: datetime
    model: str  # 'A' or 'B'
    department: str  # 'A' or 'B'
    current_month_index: int  # Index of current month in internship
    email: str = ""
    assignments: Dict[int, str] = field(default_factory=dict)  # month_index -> station_key
    total_months: int = 72
    unpaid_leave_months: int = 0
    maternity_leave_months: int = 0
    sick_leave_months_by_year: Dict[int, int] = field(default_factory=dict)  # year -> count
    
    def __post_init__(self):
        if self.model == 'B':
            self.total_months = 66
    
    def calculate_leave_counts(self):
        """Calculate leave counts from assignments."""
        self.maternity_leave_months = sum(1 for s in self.assignments.values() if s == 'חל"ד')
        self.unpaid_leave_months = sum(1 for s in self.assignments.values() if s == 'חל"ת')
        
        # Calculate sick leave by year
        self.sick_leave_months_by_year = {}
        for month_idx, station_key in self.assignments.items():
            if station_key == 'מחלה':
                month_date = self.start_date + timedelta(days=30 * month_idx)
                year = month_date.year
                self.sick_leave_months_by_year[year] = self.sick_leave_months_by_year.get(year, 0) + 1
    
    def get_effective_department_months(self) -> int:
        """
        Calculate effective department months including leave credits.
        - Maternity: Counts as dept time, capped at 6 months total
        - Sick Leave: Counts as dept time, capped at 1 month per calendar year
        - Unpaid: Does NOT count as dept time
        """
        actual_dept = sum(1 for s in self.assignments.values() if s == 'מחלקה')
        maternity_credit = min(self.maternity_leave_months, 6)
        sick_credit = sum(min(count, 1) for count in self.sick_leave_months_by_year.values())
        
        return actual_dept + maternity_credit + sick_credit
    
    def get_expected_total_months(self) -> int:
        """
        Calculate expected total program months including extensions.
        - Base: 72 (Model A) or 66 (Model B)
        - Maternity: If > 6 months, extend by (total - 6)
        - Sick Leave: Extend by any excess beyond 1 month/year
        - Unpaid: STRICTLY extend by total unpaid months
        """
        base = 72 if self.model == 'A' else 66
        
        # Maternity extension
        maternity_extension = max(0, self.maternity_leave_months - 6)
        
        # Sick leave extension
        sick_extension = sum(max(0, count - 1) for count in self.sick_leave_months_by_year.values())
        
        # Unpaid extension
        unpaid_extension = self.unpaid_leave_months
        
        return base + maternity_extension + sick_extension + unpaid_extension
    
    def get_expected_end_date(self) -> datetime:
        """
        Calculate expected end date based on start_date and program duration.
        Takes into account individual start_date and leave extensions.
        """
        expected_months = self.get_expected_total_months()
        # Add months to start_date (approximate: 30 days per month)
        return self.start_date + timedelta(days=30 * expected_months)
    
    def get_month_date(self, month_index: int) -> datetime:
        """
        Get the actual calendar date for a given month_index.
        month_index is relative to this intern's start_date.
        """
        return self.start_date + timedelta(days=30 * month_index)
    
    def calculate_progress(self) -> Dict[str, float]:
        """
        Calculate effective progress with intelligent leave handling.
        ONLY counts COMPLETED months (month_idx <= current_month_index).
        
        Rules:
        1. Base denominator: 72 (Model A) or 66 (Model B)
        2. Valid stations count as 1 each
        3. Maternity Leave (חל"ד): Counts, but capped at 6 months total
        4. Sick Leave (מחלה): Counts, but capped at 1 month per calendar year
        5. Unpaid Leave (חל"ת): Does NOT count (0)
        6. Future months (month_idx > current_month_index): Do NOT count
        
        Returns:
            Dict with 'completed', 'total', 'percent' keys
        """
        if not self.assignments:
            return {'completed': 0.0, 'total': self.total_months, 'percent': 0.0}
        
        # Count valid COMPLETED assignments only (past + current month)
        valid_months = 0
        maternity_count = 0
        sick_by_year = {}
        
        for month_idx, station_key in self.assignments.items():
            # CRITICAL: Only count months that have actually happened
            if month_idx > self.current_month_index:
                continue
            
            # Skip unpaid leave completely
            if station_key == 'חל"ת':
                continue
            
            # Track maternity leave separately (will cap at 6)
            elif station_key == 'חל"ד':
                maternity_count += 1
            
            # Track sick leave by year (will cap at 1/year)
            elif station_key == 'מחלה':
                month_date = self.get_month_date(month_idx)
                year = month_date.year
                sick_by_year[year] = sick_by_year.get(year, 0) + 1
            
            # All other stations count as 1
            else:
                valid_months += 1
        
        # Apply maternity cap (max 6 months)
        maternity_credit = min(maternity_count, 6)
        
        # Apply sick leave cap (max 1 per year)
        sick_credit = sum(min(count, 1) for count in sick_by_year.values())
        
        # Total effective completed months
        completed = valid_months + maternity_credit + sick_credit
        
        # Base program duration
        total = self.total_months  # 72 or 66
        
        # Calculate percentage
        percent = (completed / total * 100) if total > 0 else 0.0
        
        return {
            'completed': float(completed),
            'total': int(total),
            'percent': round(percent, 1)
        }


class ExcelParser:
    
    def __init__(self):
        """Initialize Excel parser."""
        pass
    
    def parse_excel(self, file_path: str, current_date: datetime) -> List[Intern]:
        """
        Parse Excel file with individual intern timelines.
        
        Structure:
        Column 1: שנה (Year)
        Column 2: חודש (Month)
        Remaining columns: Intern names (header row)
        Cell values: Hebrew station names
        
        CRITICAL: Each intern's start_date is the FIRST row where they have a non-empty cell.
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Step 1: Read header row to get intern names (DO NOT create Intern objects yet)
        header_row = ws[1]
        intern_columns = {}  # {intern_name: column_index}
        
        for col_idx, cell in enumerate(header_row, start=1):
            if col_idx <= 2:  # Skip 'שנה' and 'חודש' columns
                continue
            if cell.value and str(cell.value).strip():
                intern_name = str(cell.value).strip()
                intern_columns[intern_name] = col_idx
        
        print(f"Found {len(intern_columns)} intern columns: {list(intern_columns.keys())}")
        
        # Step 2: Dictionary to track interns (will be created on-demand)
        interns_map = {}  # {intern_name: Intern object}
        
        # Step 3: Iterate through rows chronologically
        for row_idx in range(2, ws.max_row + 1):
            # Get year and month from first two columns
            year_cell = ws.cell(row=row_idx, column=1)
            month_cell = ws.cell(row=row_idx, column=2)
            
            if not year_cell.value or not month_cell.value:
                continue  # Skip empty rows
            
            try:
                year = int(year_cell.value)
                month = int(month_cell.value)
                current_row_date = datetime(year, month, 1)
            except (ValueError, TypeError):
                print(f"Warning: Invalid date at row {row_idx}: year={year_cell.value}, month={month_cell.value}")
                continue
            
            # Step 4: For each intern column, check the cell
            for intern_name, col_idx in intern_columns.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Case A: Cell is EMPTY - skip (do NOT create intern)
                if not cell.value or not str(cell.value).strip():
                    continue
                
                # Case B: Cell has TEXT (station name)
                station_value = str(cell.value).strip()
                
                # Normalize Hebrew station name (no translation)
                station_key = normalize_station_name(station_value)
                
                # Validate against known stations
                if station_key not in config.STATIONS_MODEL_A and station_key not in config.STATIONS_MODEL_B:
                    print(f"Warning: Unknown station '{station_value}' (normalized: '{station_key}') for {intern_name} at {current_row_date.strftime('%Y-%m')}. Using 'מחלקה' as fallback.")
                    station_key = 'מחלקה'
                
                # Check if intern exists
                if intern_name not in interns_map:
                    # This is their FIRST assignment - CREATE intern with start_date = current_row_date
                    print(f"Creating {intern_name} with start_date = {current_row_date.strftime('%Y-%m-%d')}")
                    
                    # Get metadata (model, department, email)
                    model = self._get_intern_model_from_metadata(ws, intern_name)
                    department = self._get_intern_department_from_metadata(ws, intern_name)
                    email = self._get_intern_email_from_metadata(ws, intern_name)
                    
                    total_months = 72 if model == 'A' else 66
                    
                    intern = Intern(
                        name=intern_name,
                        start_date=current_row_date,  # CRITICAL: Set to THIS row's date
                        model=model,
                        department=department,
                        current_month_index=0,  # Will be calculated later
                        email=email,
                        assignments={},
                        total_months=total_months
                    )
                    
                    interns_map[intern_name] = intern
                
                # Add assignment to this intern
                intern = interns_map[intern_name]
                
                # Calculate month_index relative to THIS intern's start_date
                month_diff = (current_row_date.year - intern.start_date.year) * 12 + (current_row_date.month - intern.start_date.month)
                intern.assignments[month_diff] = station_key
        
        wb.close()
        
        # Step 5: Post-process all interns
        interns = []
        for intern_name, intern in interns_map.items():
            # Calculate current_month_index
            current_month_index = 0
            for month_idx in sorted(intern.assignments.keys()):
                month_date = intern.get_month_date(month_idx)
                if month_date <= current_date:
                    current_month_index = month_idx
            
            intern.current_month_index = current_month_index
            
            # Calculate leave counts
            intern.calculate_leave_counts()
            
            print(f"{intern.name}: start_date={intern.start_date.strftime('%Y-%m-%d')}, {len(intern.assignments)} months assigned, current_month_index={current_month_index}")
            
            interns.append(intern)
        
        print(f"Successfully parsed {len(interns)} interns with individual start dates")
        return interns
    
    def _get_intern_model_from_metadata(self, ws, intern_name: str) -> str:
        """Extract intern model (A or B) from metadata area."""
        # Look for metadata rows at the bottom of the sheet
        # Expected format: Row with "Model" or "מודל" in first column
        # Then intern names in header positions with their model values
        try:
            for row in range(ws.max_row - 20, ws.max_row + 1):
                label_cell = ws.cell(row=row, column=1)
                if label_cell.value:
                    label = str(label_cell.value).strip().lower()
                    if 'model' in label or 'מודל' in label:
                        # Find intern's column
                        for col in range(3, ws.max_column + 1):
                            header_cell = ws.cell(row=1, column=col)
                            if header_cell.value and str(header_cell.value).strip() == intern_name:
                                value_cell = ws.cell(row=row, column=col)
                                if value_cell.value:
                                    val = str(value_cell.value).strip().upper()
                                    if 'B' in val:
                                        return 'B'
                                    return 'A'
        except Exception as e:
            print(f"Warning: Could not read model metadata for {intern_name}: {e}")
        
        return 'A'  # Default to Model A
    
    def _get_intern_department_from_metadata(self, ws, intern_name: str) -> str:
        """Extract intern department (A or B) from metadata area."""
        try:
            for row in range(ws.max_row - 20, ws.max_row + 1):
                label_cell = ws.cell(row=row, column=1)
                if label_cell.value:
                    label = str(label_cell.value).strip().lower()
                    if 'department' in label or 'מחלקה' in label:
                        # Find intern's column
                        for col in range(3, ws.max_column + 1):
                            header_cell = ws.cell(row=1, column=col)
                            if header_cell.value and str(header_cell.value).strip() == intern_name:
                                value_cell = ws.cell(row=row, column=col)
                                if value_cell.value:
                                    val = str(value_cell.value).strip().upper()
                                    if 'B' in val:
                                        return 'B'
                                    return 'A'
        except Exception as e:
            print(f"Warning: Could not read department metadata for {intern_name}: {e}")
        
        return 'A'  # Default to Department A
    
    def _get_intern_email_from_metadata(self, ws, intern_name: str) -> str:
        """Extract intern email from metadata area."""
        try:
            for row in range(ws.max_row - 20, ws.max_row + 1):
                label_cell = ws.cell(row=row, column=1)
                if label_cell.value:
                    label = str(label_cell.value).strip().lower()
                    if 'email' in label or 'דוא' in label:
                        # Find intern's column
                        for col in range(3, ws.max_column + 1):
                            header_cell = ws.cell(row=1, column=col)
                            if header_cell.value and str(header_cell.value).strip() == intern_name:
                                value_cell = ws.cell(row=row, column=col)
                                if value_cell.value:
                                    return str(value_cell.value).strip()
        except Exception as e:
            print(f"Warning: Could not read email metadata for {intern_name}: {e}")
        
        return ""  # Default to empty string
    


class ExcelWriter:
    
    def write_schedule(self, interns: List[Intern], output_path: str, start_month: datetime = None):
        """
        Write intern schedules to Excel with Hebrew format:
        Column 1: שנה (Year)
        Column 2: חודש (Month)
        Remaining columns: Intern names with station assignments
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Internship Schedule"
        
        # Write header row
        ws.cell(row=1, column=1, value="שנה")
        ws.cell(row=1, column=2, value="חודש")
        for col_idx, intern in enumerate(interns, start=3):
            cell = ws.cell(row=1, column=col_idx, value=intern.name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Determine date range
        if start_month is None:
            start_month = min(intern.start_date for intern in interns)
        
        # Collect all dates across all interns
        all_dates = set()
        for intern in interns:
            for month_idx in intern.assignments.keys():
                date = intern.start_date + timedelta(days=30 * month_idx)
                all_dates.add((date.year, date.month))
        
        # Sort dates
        sorted_dates = sorted(all_dates)
        
        # Write data rows
        for row_idx, (year, month) in enumerate(sorted_dates, start=2):
            ws.cell(row=row_idx, column=1, value=year)
            ws.cell(row=row_idx, column=2, value=month)
            
            row_date = datetime(year, month, 1)
            
            # Write assignments for each intern
            for col_idx, intern in enumerate(interns, start=3):
                # Calculate month_index for this intern
                month_diff = (row_date.year - intern.start_date.year) * 12 + (row_date.month - intern.start_date.month)
                
                if month_diff in intern.assignments:
                    station_key = intern.assignments[month_diff]
                    
                    # Get station info
                    stations = config.STATIONS_MODEL_A if intern.model == 'A' else config.STATIONS_MODEL_B
                    station = stations.get(station_key)
                    
                    if station:
                        cell = ws.cell(row=row_idx, column=col_idx, value=station.name)
                        
                        # Apply color
                        fill = PatternFill(start_color=station.color.lstrip('#'), 
                                         end_color=station.color.lstrip('#'), 
                                         fill_type='solid')
                        cell.fill = fill
                        cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for col in range(1, len(interns) + 3):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Add metadata section at the bottom
        metadata_start_row = len(sorted_dates) + 5
        ws.cell(row=metadata_start_row, column=1, value="Start Date")
        ws.cell(row=metadata_start_row + 1, column=1, value="Model")
        ws.cell(row=metadata_start_row + 2, column=1, value="Department")
        ws.cell(row=metadata_start_row + 3, column=1, value="Email")
        
        for col_idx, intern in enumerate(interns, start=3):
            ws.cell(row=metadata_start_row, column=col_idx, value=intern.start_date.strftime("%Y-%m-%d"))
            ws.cell(row=metadata_start_row + 1, column=col_idx, value=intern.model)
            ws.cell(row=metadata_start_row + 2, column=col_idx, value=intern.department)
            ws.cell(row=metadata_start_row + 3, column=col_idx, value=intern.email if intern.email else "")
        
        wb.save(output_path)
        wb.close()
    
    def create_sample_excel(self, output_path: str):
        """Create a sample Excel file with Hebrew format for testing."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sample Schedule"
        
        # Header row with Hebrew
        ws.cell(row=1, column=1, value="שנה")
        ws.cell(row=1, column=2, value="חודש")
        ws.cell(row=1, column=3, value="מתמחה 1")
        ws.cell(row=1, column=4, value="מתמחה 2")
        
        # Sample data (6 months)
        sample_data = [
            (2024, 1, "אוריינטציה", "אוריינטציה"),
            (2024, 2, "יולדות", "יולדות"),
            (2024, 3, "הריון בסיכון א", "הריון בסיכון ב"),
            (2024, 4, "הריון בסיכון א", "הריון בסיכון ב"),
            (2024, 5, "הריון בסיכון א", "הריון בסיכון ב"),
            (2024, 6, "הריון בסיכון א", "הריון בסיכון ב"),
        ]
        
        for row_idx, (year, month, station1, station2) in enumerate(sample_data, start=2):
            ws.cell(row=row_idx, column=1, value=year)
            ws.cell(row=row_idx, column=2, value=month)
            ws.cell(row=row_idx, column=3, value=station1)
            ws.cell(row=row_idx, column=4, value=station2)
        
        # Metadata section
        metadata_row = len(sample_data) + 5
        ws.cell(row=metadata_row, column=1, value="Start Date")
        ws.cell(row=metadata_row, column=3, value="2024-01-01")
        ws.cell(row=metadata_row, column=4, value="2024-01-01")
        
        ws.cell(row=metadata_row + 1, column=1, value="Model")
        ws.cell(row=metadata_row + 1, column=3, value="A")
        ws.cell(row=metadata_row + 1, column=4, value="A")
        
        ws.cell(row=metadata_row + 2, column=1, value="Department")
        ws.cell(row=metadata_row + 2, column=3, value="A")
        ws.cell(row=metadata_row + 2, column=4, value="B")
        
        ws.cell(row=metadata_row + 3, column=1, value="Email")
        ws.cell(row=metadata_row + 3, column=3, value="intern1@example.com")
        ws.cell(row=metadata_row + 3, column=4, value="intern2@example.com")
        
        wb.save(output_path)
        wb.close()

